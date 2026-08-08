"""
FULL REPORT — Excel consolidation + Gmail email

Reads the latest JSON output from all 7 Phase 1/2 modules (whichever
are available — missing ones are skipped, not treated as errors) and
builds a single multi-sheet Excel workbook:

    Sector Performance | Paper Trading | Analysis Engine |
    Learning Engine | Optimizer | Backtest | Regression

Then emails that workbook as an attachment via Gmail SMTP (app
password — NOT your real Gmail password; generate one at
https://myaccount.google.com/apppasswords).

Required environment variables (set as GitHub Actions secrets):
    GMAIL_ADDRESS       — the Gmail account sending the email
    GMAIL_APP_PASSWORD  — 16-character app password (not your login password)
    REPORT_RECIPIENT    — where to send the report (can be the same address)

Usage:
    python scripts/email_full_report.py
"""

from __future__ import annotations

import json
import os
import smtplib
import sys
import time
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from core.logger import get_logger  # noqa: E402

logger = get_logger(__name__)

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11)
BODY_FONT = Font(name=FONT_NAME, size=10)

# (sheet title, JSON file path)
REPORT_SOURCES = [
    ("Sector Performance", "reports/sector_performance_latest.json"),
    ("Paper Trading", "reports/paper_trading_summary_latest.json"),
    ("Analysis Engine", "reports/analysis_summary.json"),
    ("Learning Engine", "reports/learning_observation_latest.json"),
    ("Optimizer", "reports/optimizer_recommendations_latest.json"),
    ("Backtest", "reports/backtest_result_latest.json"),
    ("Regression", "reports/regression_result_latest.json"),
]

OUTPUT_XLSX = "reports/full_report_latest.xlsx"


def _write_value_rows(ws: Worksheet, data: Any, row: int, indent: int = 0) -> int:
    """Recursively write a JSON-like structure as Key | Value rows.
    Nested dicts get indented sub-rows; lists of dicts become a small
    inline table (one row per item, one column per key found)."""
    prefix = "    " * indent
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, (dict, list)):
                ws.cell(row=row, column=1, value=f"{prefix}{key}").font = HEADER_FONT if indent == 0 else BODY_FONT
                row += 1
                row = _write_value_rows(ws, value, row, indent + 1)
            else:
                ws.cell(row=row, column=1, value=f"{prefix}{key}").font = BODY_FONT
                ws.cell(row=row, column=2, value=value).font = BODY_FONT
                row += 1
    elif isinstance(data, list):
        if not data:
            ws.cell(row=row, column=1, value=f"{prefix}(empty)").font = BODY_FONT
            row += 1
        elif all(isinstance(item, dict) for item in data):
            # Table: union of all keys across items, as columns
            keys: list[str] = []
            for item in data:
                for k in item.keys():
                    if k not in keys:
                        keys.append(k)
            for col, k in enumerate(keys, start=2):
                ws.cell(row=row, column=col, value=k).font = HEADER_FONT
            row += 1
            for item in data:
                for col, k in enumerate(keys, start=2):
                    ws.cell(row=row, column=col, value=item.get(k)).font = BODY_FONT
                row += 1
        else:
            for item in data:
                ws.cell(row=row, column=1, value=f"{prefix}- {item}").font = BODY_FONT
                row += 1
    else:
        ws.cell(row=row, column=1, value=f"{prefix}{data}").font = BODY_FONT
        row += 1
    return row


def _build_sheet(wb: Workbook, title: str, json_path: str) -> None:
    ws = wb.create_sheet(title=title[:31])  # Excel sheet name limit
    ws.column_dimensions["A"].width = 34
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 20

    if not Path(json_path).exists():
        ws.cell(row=1, column=1, value=f"No data available yet — {json_path} not found.").font = BODY_FONT
        return

    try:
        with open(json_path) as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as exc:
        ws.cell(row=1, column=1, value=f"Could not read {json_path}: {exc}").font = BODY_FONT
        return

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name=FONT_NAME, bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=2, column=1, value=f"Source: {json_path}").font = Font(name=FONT_NAME, italic=True, size=9)

    _write_value_rows(ws, data, row=4)


def build_workbook() -> str:
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for title, json_path in REPORT_SOURCES:
        _build_sheet(wb, title, json_path)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No module reports were available yet.")

    Path("reports").mkdir(exist_ok=True)
    wb.save(OUTPUT_XLSX)
    logger.info("Full report workbook written to %s", OUTPUT_XLSX)
    return OUTPUT_XLSX


def send_email(xlsx_path: str) -> None:
    gmail_address = os.environ.get("GMAIL_ADDRESS")
    gmail_app_password = os.environ.get("GMAIL_APP_PASSWORD")
    recipient = os.environ.get("REPORT_RECIPIENT")

    if not gmail_address or not gmail_app_password or not recipient:
        logger.warning(
            "GMAIL_ADDRESS / GMAIL_APP_PASSWORD / REPORT_RECIPIENT not all set — "
            "workbook was built (%s) but NOT emailed.", xlsx_path,
        )
        print(f"Workbook built at {xlsx_path}, but email env vars are missing — skipped sending.")
        return

    msg = MIMEMultipart()
    msg["From"] = gmail_address
    msg["To"] = recipient
    msg["Subject"] = f"Full Trading System Report — {time.strftime('%Y-%m-%d')}"
    msg.attach(MIMEText(
        "Attached: consolidated report covering Sector Performance, Paper Trading, "
        "Analysis Engine, Learning Engine, Optimizer, Backtest, and Regression.\n\n"
        "This is an automated email.", "plain",
    ))

    with open(xlsx_path, "rb") as f:
        part = MIMEApplication(f.read(), _subtype="xlsx")
    part.add_header("Content-Disposition", "attachment", filename=Path(xlsx_path).name)
    msg.attach(part)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(gmail_address, gmail_app_password)
            server.send_message(msg)
        logger.info("Full report emailed to %s", recipient)
        print(f"Emailed {xlsx_path} to {recipient}.")
    except smtplib.SMTPException as exc:
        logger.warning("Failed to send full report email: %s", exc)
        print(f"Email failed: {exc}")


def main() -> None:
    xlsx_path = build_workbook()
    send_email(xlsx_path)


if __name__ == "__main__":
    main()
